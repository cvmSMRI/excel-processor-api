from flask import Flask, request, send_file
import pandas as pd
import openpyxl
import io

app = Flask(__name__)

# List your required target columns for each file
KSP_COLUMNS = ["SKU No.", "Item Description", "Vendor Name", "Unit Cost", "Category", "Status"]
TWP_COLUMNS = ["SKU No.", "Selling Price", "Stock Qty", "Branch Code", "Reorder Level"]

@app.route('/process-template', methods=['POST'])
def process_template():
    try:
        # Check that both CSV files and the Template are passed in the request
        if 'ksp_file' not in request.files or 'twp_file' not in request.files or 'template_file' not in request.files:
            return {"error": "Missing required files (ksp_file, twp_file, or template_file)"}, 400
        
        ksp_csv = request.files['ksp_file']
        twp_csv = request.files['twp_file']
        template_excel = request.files['template_file']

        # 1. Read CSV files using Pandas
        df_ksp = pd.read_csv(ksp_csv)
        df_twp = pd.read_csv(twp_csv)

        # 2. Filter target columns
        ksp_filtered = df_ksp[[col for col in KSP_COLUMNS if col in df_ksp.columns]]
        twp_filtered = df_twp[[col for col in TWP_COLUMNS if col in df_twp.columns]]

        # 3. Load the SharePoint Excel template into OpenPyXL/Pandas writer
        output_stream = io.BytesIO()
        
        with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
            # Load existing worksheets from your template
            writer.workbook = openpyxl.load_workbook(template_excel)
            
            # Write data into the respective sheet tabs
            if 'KSP' in writer.workbook.sheetnames:
                ksp_filtered.to_excel(writer, sheet_name='KSP', index=False)
            if 'TWP' in writer.workbook.sheetnames:
                twp_filtered.to_excel(writer, sheet_name='TWP', index=False)

        output_stream.seek(0)

        # 4. Return the populated Excel template file
        return send_file(
            output_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Final_Populated_Report.xlsx'
        )

    except Exception as e:
        return {"status": "error", "message": str(e)}, 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)
